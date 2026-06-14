"""Per-format handler functions for ``Spectrum`` load/save (FR-132..FR-134).

Each function corresponds to one ``handler=`` named in a ``LoadSpectrum`` /
``SaveSpectrum`` ``FormatCapability`` record (spec §"Spectrum load/save
capabilities"). The block methods delegate here.

Loaders return a single :class:`Spectrum` built via
``_support.build_spectrum`` (``spectrum_id`` left ``None`` so a fresh package
id is generated — never derived from the filename, FR-035/FR-036) with
``source_file=str(path)`` recorded as metadata only. Savers write *spectrum* to
*path* and return ``None``. Heavy/optional parsers (openpyxl, vendor SDKs) are
lazy-imported inside the function body, never at module top level.

Implemented round-trippable formats:

- ``delimited_text`` (``.txt``/``.csv``/``.tsv``) — two-column lambda,intensity;
  delimiter sniffed from suffix; comment/header lines skipped. ``pixel_only``.
- ``spectrum_xlsx`` (``.xlsx``) — two-column data sheet plus a ``meta`` sheet for
  the typed ``Spectrum.Meta`` fields. ``typed_meta``.
- ``spectrum_json`` (``.spectrum.json``) — package-owned native LOSSLESS JSON.
- ``jcamp_dx`` (``.jdx``/``.dx``/``.jcamp``) — minimal JCAMP-DX (``##XYPOINTS``)
  with ``##XUNITS``/``##YUNITS`` mapped to lambda/intensity units. ``typed_meta``.

``spc`` and the vendor LOAD-ONLY formats raise an informative
``NotImplementedError`` (no fixture / optional SDK), tracked under #1661.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

import numpy as np

from scistudio_blocks_spectroscopy import _support
from scistudio_blocks_spectroscopy.types import (
    INTENSITY_COLUMN,
    LAMBDA_COLUMN,
    Spectrum,
)

# Typed ``Spectrum.Meta`` fields carried by the typed_meta / lossless formats.
_TYPED_META_FIELDS = ("lambda_unit", "intensity_unit", "lambda_kind", "modality")
# Native JSON schema marker so a future change can branch on the on-disk shape.
_SPECTRUM_JSON_SCHEMA = "scistudio-blocks-spectroscopy/spectrum-json/1"


# --------------------------------------------------------------------------
# Shared helpers
# --------------------------------------------------------------------------


def _delimiter_for_suffix(path: Path) -> str:
    """Return the column delimiter implied by *path*'s suffix.

    ``.csv`` -> comma, ``.tsv`` -> tab, everything else (``.txt``) -> whitespace
    runs (represented by ``None`` for :func:`str.split`).
    """
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return ","
    if suffix == ".tsv":
        return "\t"
    return " "


def _split_row(line: str, delimiter: str) -> list[str]:
    """Split *line* into fields using *delimiter* (whitespace when ``" "``)."""
    if delimiter == " ":
        return line.split()
    return [field.strip() for field in line.split(delimiter)]


def _typed_meta_kwargs(spectrum: Spectrum) -> dict[str, Any]:
    """Extract the typed ``Spectrum.Meta`` fields as a plain dict."""
    meta = spectrum.meta
    if not isinstance(meta, Spectrum.Meta):
        return {field: None for field in _TYPED_META_FIELDS}
    return {field: getattr(meta, field) for field in _TYPED_META_FIELDS}


# --------------------------------------------------------------------------
# Delimited text (.txt / .csv / .tsv) — load + save, pixel_only
# --------------------------------------------------------------------------


def load_delimited_text(path: Path, **kwargs: Any) -> Spectrum:
    """Load a ``.txt``/``.csv``/``.tsv`` two-column spectrum (FR-132, FR-142).

    Sniffs the delimiter from the suffix, skips comment lines (``#``/``%``/
    ``;``) and any non-numeric header rows, then reads the first two numeric
    columns into ``(lambda, intensity)``. ``metadata_fidelity`` is
    ``pixel_only`` so no typed meta is read.
    """
    path = Path(path)
    delimiter = _delimiter_for_suffix(path)
    lambdas: list[float] = []
    intensities: list[float] = []
    text = path.read_text(encoding="utf-8-sig")
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line or line[0] in "#%;":
            continue
        fields = _split_row(line, delimiter)
        if len(fields) < 2:
            # Single-column rows are not a valid two-column spectrum.
            raise ValueError(f"load_delimited_text: expected two numeric columns per row in {path}, got {fields!r}")
        try:
            lam = float(fields[0])
            inten = float(fields[1])
        except ValueError:
            # Header / units row with non-numeric cells: skip it.
            continue
        lambdas.append(lam)
        intensities.append(inten)
    if not lambdas:
        raise ValueError(f"load_delimited_text: no numeric two-column rows found in {path}")
    return _support.build_spectrum(
        np.asarray(lambdas, dtype=np.float64),
        np.asarray(intensities, dtype=np.float64),
        meta=Spectrum.Meta(source_file=str(path)),
    )


def save_delimited_text(spectrum: Spectrum, path: Path, **kwargs: Any) -> None:
    """Save a spectrum as ``.txt``/``.csv``/``.tsv`` two columns (FR-132, FR-037).

    Picks the delimiter from the suffix. ``metadata_fidelity`` is ``pixel_only``
    so typed meta is intentionally not written.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    lam, inten = _support.spectrum_arrays(spectrum)
    delimiter = _delimiter_for_suffix(path)
    if delimiter == " ":
        lines = [f"{x:.10g} {y:.10g}" for x, y in zip(lam, inten, strict=True)]
        path.write_text("\n".join(lines) + ("\n" if lines else ""), encoding="utf-8")
        return
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle, delimiter=delimiter)
        for x, y in zip(lam, inten, strict=True):
            writer.writerow([f"{x:.10g}", f"{y:.10g}"])


# --------------------------------------------------------------------------
# Excel workbook (.xlsx) — load + save, typed_meta
# --------------------------------------------------------------------------


def _require_openpyxl() -> Any:
    """Lazy-import openpyxl, raising a clear install hint when missing."""
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - exercised only without extra
        raise ImportError(
            "Excel (.xlsx) spectrum IO requires 'openpyxl'. Install it with "
            "`pip install openpyxl` (or the package's [excel] extra)."
        ) from exc
    return openpyxl


def load_spectrum_xlsx(path: Path, **kwargs: Any) -> Spectrum:
    """Load a single spectrum from an ``.xlsx``/``.xls`` workbook (FR-132).

    Reads the first sheet's two ``lambda``/``intensity`` columns (header row
    skipped) plus an optional ``meta`` sheet of ``key``/``value`` pairs that
    carries the typed ``Spectrum.Meta`` fields. ``typed_meta`` fidelity.
    """
    path = Path(path)
    openpyxl = _require_openpyxl()
    workbook = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        data_sheet = workbook["data"] if "data" in workbook.sheetnames else workbook.worksheets[0]
        lambdas: list[float] = []
        intensities: list[float] = []
        for row in data_sheet.iter_rows(values_only=True):
            if row is None or len(row) < 2 or row[0] is None or row[1] is None:
                continue
            try:
                lam = float(row[0])
                inten = float(row[1])
            except (TypeError, ValueError):
                # Header / label row.
                continue
            lambdas.append(lam)
            intensities.append(inten)

        meta_kwargs: dict[str, Any] = {}
        if "meta" in workbook.sheetnames:
            for row in workbook["meta"].iter_rows(values_only=True):
                if not row or row[0] is None:
                    continue
                key = str(row[0]).strip()
                value = row[1] if len(row) > 1 else None
                if key in _TYPED_META_FIELDS:
                    meta_kwargs[key] = None if value in (None, "") else str(value)
    finally:
        workbook.close()

    if not lambdas:
        raise ValueError(f"load_spectrum_xlsx: no numeric two-column rows found in {path}")
    return _support.build_spectrum(
        np.asarray(lambdas, dtype=np.float64),
        np.asarray(intensities, dtype=np.float64),
        meta=Spectrum.Meta(source_file=str(path), **meta_kwargs),
    )


def save_spectrum_xlsx(spectrum: Spectrum, path: Path, **kwargs: Any) -> None:
    """Save a single spectrum to an ``.xlsx`` workbook with typed meta (FR-132).

    Writes a ``data`` sheet (``lambda``/``intensity`` columns with a header
    row) plus a ``meta`` sheet of ``key``/``value`` pairs for the typed
    ``Spectrum.Meta`` fields — ``typed_meta`` fidelity.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    openpyxl = _require_openpyxl()
    lam, inten = _support.spectrum_arrays(spectrum)

    workbook = openpyxl.Workbook()
    data_sheet = workbook.active
    data_sheet.title = "data"
    data_sheet.append([LAMBDA_COLUMN, INTENSITY_COLUMN])
    for x, y in zip(lam, inten, strict=True):
        data_sheet.append([float(x), float(y)])

    meta_sheet = workbook.create_sheet("meta")
    meta_sheet.append(["key", "value"])
    for field, value in _typed_meta_kwargs(spectrum).items():
        meta_sheet.append([field, "" if value is None else str(value)])

    workbook.save(str(path))


# --------------------------------------------------------------------------
# Native Spectrum JSON (.spectrum.json) — load + save, LOSSLESS
# --------------------------------------------------------------------------


def load_spectrum_json(path: Path, **kwargs: Any) -> Spectrum:
    """Load the package-native ``.spectrum.json`` format (FR-132, FR-141).

    Lossless round-trip: reads ``lambda``/``intensity`` arrays plus the full
    typed ``Spectrum.Meta`` (including ``spectrum_id``) and the free-form
    ``user`` dict.
    """
    path = Path(path)
    payload = json.loads(path.read_text(encoding="utf-8"))
    lam = payload.get(LAMBDA_COLUMN)
    inten = payload.get(INTENSITY_COLUMN)
    if lam is None or inten is None:
        raise ValueError(f"load_spectrum_json: missing 'lambda'/'intensity' arrays in {path}")

    raw_meta = payload.get("meta") or {}
    # Only feed known Meta fields to the model; unknown keys would be rejected
    # by the frozen pydantic model. Lossless round-trip preserves every field
    # the native saver emits (all of Spectrum.Meta).
    valid_fields = set(Spectrum.Meta.model_fields)
    meta_kwargs = {key: value for key, value in raw_meta.items() if key in valid_fields}
    user = payload.get("user") or None

    return _support.build_spectrum(
        np.asarray(lam, dtype=np.float64),
        np.asarray(inten, dtype=np.float64),
        meta=Spectrum.Meta(**meta_kwargs),
        user=user,
    )


def save_spectrum_json(spectrum: Spectrum, path: Path, **kwargs: Any) -> None:
    """Save the package-native ``.spectrum.json`` format (FR-132, FR-141).

    Lossless: serialises ``lambda``/``intensity`` plus the entire typed
    ``Spectrum.Meta`` and the ``user`` dict.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    lam, inten = _support.spectrum_arrays(spectrum)

    meta = spectrum.meta
    meta_dict = meta.model_dump(mode="json") if isinstance(meta, Spectrum.Meta) else {}

    document = {
        "schema": _SPECTRUM_JSON_SCHEMA,
        LAMBDA_COLUMN: [float(x) for x in lam],
        INTENSITY_COLUMN: [float(y) for y in inten],
        "meta": meta_dict,
        "user": dict(spectrum.user) if spectrum.user else {},
    }
    path.write_text(json.dumps(document, indent=2, sort_keys=False), encoding="utf-8")


# --------------------------------------------------------------------------
# JCAMP-DX (.jdx / .dx / .jcamp) — load + save, typed_meta
# --------------------------------------------------------------------------


def load_jcamp_dx(path: Path, **kwargs: Any) -> Spectrum:
    """Load a JCAMP-DX (``.jdx``/``.dx``/``.jcamp``) spectrum (FR-132).

    Parses the minimal AFFN ``##XYPOINTS`` / ``##XYDATA=(XY..XY)`` layout this
    package writes: ``(x, y)`` pairs one per line. Maps ``##XUNITS``/
    ``##YUNITS`` to ``lambda_unit``/``intensity_unit`` (``typed_meta``).
    """
    path = Path(path)
    x_units: str | None = None
    y_units: str | None = None
    lambdas: list[float] = []
    intensities: list[float] = []
    in_data = False

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("$$"):
            continue
        if line.startswith("##"):
            label, _, value = line[2:].partition("=")
            label = label.strip().upper()
            value = value.strip()
            if label == "XUNITS":
                x_units = value or None
            elif label == "YUNITS":
                y_units = value or None
            elif label in ("XYDATA", "XYPOINTS", "PEAK TABLE"):
                in_data = True
            elif label == "END":
                in_data = False
            continue
        if not in_data:
            continue
        # Data line: a list of (x, y) pairs separated by ';' or whitespace,
        # each pair joined by ',' or whitespace.
        for token in line.replace(";", " ").split():
            x_str, sep, y_str = token.partition(",")
            if not sep:
                continue
            try:
                lambdas.append(float(x_str))
                intensities.append(float(y_str))
            except ValueError:
                continue

    if not lambdas:
        raise ValueError(f"load_jcamp_dx: no XY data points found in {path}")
    return _support.build_spectrum(
        np.asarray(lambdas, dtype=np.float64),
        np.asarray(intensities, dtype=np.float64),
        meta=Spectrum.Meta(source_file=str(path), lambda_unit=x_units, intensity_unit=y_units),
    )


def save_jcamp_dx(spectrum: Spectrum, path: Path, **kwargs: Any) -> None:
    """Save a spectrum as JCAMP-DX (``.jdx``/``.dx``/``.jcamp``) (FR-132).

    Emits the required ``##TITLE``/``##JCAMP-DX``/``##DATA TYPE`` headers, maps
    ``lambda_unit``/``intensity_unit`` to ``##XUNITS``/``##YUNITS``, and writes
    an AFFN ``##XYPOINTS=(XY..XY)`` block of ``x,y`` pairs.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    lam, inten = _support.spectrum_arrays(spectrum)
    typed = _typed_meta_kwargs(spectrum)
    x_units = typed.get("lambda_unit") or "ARBITRARY UNITS"
    y_units = typed.get("intensity_unit") or "ARBITRARY UNITS"

    title = "Spectrum"
    meta = spectrum.meta
    if isinstance(meta, Spectrum.Meta) and meta.sample_label:
        title = meta.sample_label

    lines = [
        f"##TITLE={title}",
        "##JCAMP-DX=4.24",
        "##DATA TYPE=SPECTRUM",
        f"##XUNITS={x_units}",
        f"##YUNITS={y_units}",
        f"##NPOINTS={len(lam)}",
        f"##FIRSTX={lam[0]:.10g}" if len(lam) else "##FIRSTX=0",
        f"##LASTX={lam[-1]:.10g}" if len(lam) else "##LASTX=0",
        "##XYPOINTS=(XY..XY)",
    ]
    lines.extend(f"{x:.10g},{y:.10g}" for x, y in zip(lam, inten, strict=True))
    lines.append("##END=")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


# --------------------------------------------------------------------------
# SPC (.spc) — deferred (needs fixture / spc library)
# --------------------------------------------------------------------------


def load_spc(path: Path, **kwargs: Any) -> Spectrum:
    """Load a Galactic/Thermo SPC (``.spc``) single spectrum (FR-132, FR-138)."""
    # TODO(#1661): SPC binary format needs a fixture/spc lib.
    #   A correct Galactic SPC reader/writer requires the documented binary
    #   header layout and a real fixture to validate against; neither is
    #   available without an optional dependency.
    #   Followup: https://github.com/zjzcpj/SciStudio/issues/1661
    raise NotImplementedError(
        "SPC (.spc) load is not implemented: the Galactic SPC binary format needs a "
        "fixture/spc library to implement correctly (TODO #1661)."
    )


def save_spc(spectrum: Spectrum, path: Path, **kwargs: Any) -> None:
    """Save a spectrum as SPC (``.spc``) single subfile (FR-132, FR-138)."""
    # TODO(#1661): SPC binary format needs a fixture/spc lib.
    #   Writing a spec-correct SPC subfile + header is infeasible without the
    #   documented binary layout and a round-trip fixture to validate.
    #   Followup: https://github.com/zjzcpj/SciStudio/issues/1661
    raise NotImplementedError(
        "SPC (.spc) save is not implemented: the Galactic SPC binary format needs a "
        "fixture/spc library to implement correctly (TODO #1661)."
    )


# --------------------------------------------------------------------------
# Vendor / instrument-native LOAD-ONLY formats (FR-133, FR-140) — deferred
# --------------------------------------------------------------------------

_VENDOR_NO_FIXTURE = "vendor binary format — no fixture/optional SDK"


def load_thermo_omnic_spa(path: Path, **kwargs: Any) -> Spectrum:
    """Load a Thermo OMNIC ``.spa`` single spectrum (load-only, FR-133)."""
    # TODO(#1661): Thermo OMNIC SPA binary format needs a fixture/optional SDK.
    #   Followup: https://github.com/zjzcpj/SciStudio/issues/1661
    raise NotImplementedError(_VENDOR_NO_FIXTURE)


def load_bruker_opus(path: Path, **kwargs: Any) -> Spectrum:
    """Load a Bruker OPUS (``.opus``) single spectrum (load-only, FR-133)."""
    # TODO(#1661): Bruker OPUS binary format needs a fixture/optional SDK.
    #   Followup: https://github.com/zjzcpj/SciStudio/issues/1661
    raise NotImplementedError(_VENDOR_NO_FIXTURE)


def load_horiba_labspec(path: Path, **kwargs: Any) -> Spectrum:
    """Load a HORIBA LabSpec single-spectrum export (load-only, FR-133)."""
    # TODO(#1661): HORIBA LabSpec binary/xml formats need a fixture/optional SDK.
    #   Followup: https://github.com/zjzcpj/SciStudio/issues/1661
    raise NotImplementedError(_VENDOR_NO_FIXTURE)


def load_renishaw_wdf(path: Path, **kwargs: Any) -> Spectrum:
    """Load a Renishaw WiRE ``.wdf`` single spectrum (load-only, FR-133)."""
    # TODO(#1661): Renishaw WDF binary format needs a fixture/optional SDK.
    #   Followup: https://github.com/zjzcpj/SciStudio/issues/1661
    raise NotImplementedError(_VENDOR_NO_FIXTURE)


def load_andor_solis(path: Path, **kwargs: Any) -> Spectrum:
    """Load an Andor Solis (``.sif``/``.fits``/``.fit``/``.asc``) spectrum (load-only, FR-133)."""
    # TODO(#1661): Andor SIF binary + FITS multi-spectrum formats need a
    #   fixture/optional SDK (astropy) to implement and validate. The trivial
    #   ASCII (.asc) case is two-column text already covered by the
    #   delimited-text capability, so it is not duplicated here.
    #   Followup: https://github.com/zjzcpj/SciStudio/issues/1661
    raise NotImplementedError(_VENDOR_NO_FIXTURE)


def load_princeton_spe(path: Path, **kwargs: Any) -> Spectrum:
    """Load a Princeton/LightField ``.spe`` single spectrum (load-only, FR-133)."""
    # TODO(#1661): Princeton SPE v2/v3 binary format needs a fixture/optional SDK.
    #   Followup: https://github.com/zjzcpj/SciStudio/issues/1661
    raise NotImplementedError(_VENDOR_NO_FIXTURE)


__all__ = [
    "load_andor_solis",
    "load_bruker_opus",
    "load_delimited_text",
    "load_horiba_labspec",
    "load_jcamp_dx",
    "load_princeton_spe",
    "load_renishaw_wdf",
    "load_spc",
    "load_spectrum_json",
    "load_spectrum_xlsx",
    "load_thermo_omnic_spa",
    "save_delimited_text",
    "save_jcamp_dx",
    "save_spc",
    "save_spectrum_json",
    "save_spectrum_xlsx",
]
